from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from general.models import  Student, Subject_Assignment, Faculty, Faculty_Subject_Assignment,Programme,Subject,Department,Subject_Category,Degree_Categaroy,Chapter
from resource.models import Post, Resource, Video
from virtual_class.models import ClassRoom
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import django_filters
import django_filters as filters
from django.contrib.auth.models import User, Group
from .forms import StudentEditForm, UserForm,TeacherEditForm,VideoForm,PostForm
from django.template.loader import render_to_string
from django.http import JsonResponse
from django.contrib.auth.hashers import make_password
import random
from django.db.models import Q, Max, Avg    
from quiz.models import Quiz,Attempt
from virtual_class.models import Assignment,Submission, Material, Forum
from faculty.models import Log_Book_Entry
from datetime import date
from tablib import Dataset
import openpyxl
year_options=('1','2','3','4','5')
year=(('','Choose Year'),('1','I - Year'),('2','II - Year'),('3','III - Year'),('4','IV - Year'),('5','V - Year'))
semester=('1','2','3','4','5','6','7','8')
section_options = ('A','B','C','D','E','F','G','H')




# Create your views here.

def is_college(user):
    return user.groups.filter(name='College').exists()


@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def college_admin(request):
    college=request.user
    category=college.college.category
    student_count = Student.objects.filter(programme__department__degree_category=category).count()
    teacher_count = Faculty.objects.filter(programme__department__degree_category=category).count()
    post_count = Post.objects.filter(degree_categaroy=category).count()
    total_user_count = student_count  + teacher_count 
    resource_count = Resource.objects.filter(subject_assignment__programme__department__degree_category=category).count()
    recent_students = Student.objects.filter(programme__department__degree_category=category).order_by('id')[:20]
    recent_teachers = Faculty.objects.filter(programme__department__degree_category=category).order_by('id')[:20]
    
    context = {
        'total_user_count':total_user_count,
        'student_count':student_count,
        'teacher_count':teacher_count,
        'students':recent_students,
        'teachers':recent_teachers,
        'post_count':post_count,
        'resource_count':resource_count,
    }

    return render(request,'admin/custom/school/school_admin.html',context)

#Student Management

class StudentFilter(django_filters.FilterSet):
    #programme = django_filters.ModelMultipleChoiceFilter(queryset=Programme.objects.filter(department__degree_category = request.user.college.category))    
    class Meta:
        model = Student
        fields = ['programme','acadmic_year','year','semester','district','gender']

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def students_list(request):
    name = request.GET.get('name')
    if name:
        students_list = Student.objects.filter(Q(user__last_name__icontains=name) | Q(user__first_name__icontains=name), programme__department__degree_category=request.user.college.category).distinct()
    else:
        students_list = Student.objects.filter(programme__department__degree_category=request.user.college.category)

    
    student_filter = StudentFilter(request.GET, queryset=students_list)

    paginator = Paginator(student_filter.qs, 20)
    page = request.GET.get('page')
    try:
        dataqs = paginator.page(page)
    except PageNotAnInteger:
        dataqs = paginator.page(1)
    except EmptyPage:
        dataqs = paginator.page(paginator.num_pages)
    
    context = {     
        'filter': student_filter,
        'name':name,
        'dataqs':dataqs,
    }
    return render(request,'admin/custom/school/students_list.html',context)

#Teacher Management

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def faculty_list(request):
    name = request.GET.get('name')
    grade = request.GET.get('grade')
    department = request.GET.get('department')
    print(department)
    department_option = Department.objects.filter(degree_category = request.user.college.category)
    grade_list = Programme.objects.all()
    teachers_list = Faculty.objects.filter(category=request.user.college.category)
    if name:
      
        teachers_list = teachers_list.filter(Q(user__last_name__icontains=name) | Q(user__first_name__icontains=name),school=request.user.school)
    if grade:
        teachers_id = Faculty_Subject_Assignment.objects.filter(subject__programme__in=grade,faculty__category=request.user.college.category).distinct().values_list('faculty__id', flat=True)
        teachers_list = teachers_list.filter(id__in=teachers_id)
    if department:
        teachers_id = Faculty_Subject_Assignment.objects.filter(faculty__department=department,faculty__category=request.user.college.category).distinct().values_list('faculty__id', flat=True)
        teachers_list = teachers_list.filter(id__in=teachers_id)
    
    context = {     
        'teachers_list': teachers_list,
        'department':department,
        'name':name,
        'grade':grade,
        'department_option':department_option,
        'grade_list':grade_list,
    }

    return render(request,'admin/custom/school/teachers_list.html',context)

#Score

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def score(request):
    programmes = Programme.objects.filter(department__degree_category =request.user.college.category)
    print(programmes)
    grade_list = []
    year_list = []
    no_of_students = []
    no_of_teachers = []
  
    for programme in programmes:
        for year_option in year_options:
            students = Student.objects.filter(year=year_option,programme=programme,programme__department__degree_category=request.user.college.category)
            if students:
                no_of_students.append(students.count())
                grade_list.append(programme)
                year_list.append(year_option)
                teachers_count =  Faculty_Subject_Assignment.objects.filter(subject__programme=programme,subject__year=year_option,subject__programme__department__degree_category=request.user.college.category).distinct().count()
                print("test-students",teachers_count)
                no_of_teachers.append(teachers_count)
    print("gradelist",grade_list,no_of_students)
    zippedList=[]
    zippedList = zip(grade_list,year_list,no_of_students,no_of_teachers)
    context = {
        'zippedList':zippedList,
        
    }
    return render(request,'admin/custom/school/score.html',context)

#Post Management

class PostFilter(django_filters.FilterSet):  
     
    class Meta:
        model = Post
        fields = ['post_title','post_category','programme','degree_categaroy']

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def view_post(request):
    title = request.GET.get('title')
    post_list = Post.objects.filter(posted_by=request.user).order_by('-id')
    if title:
        post_list = post_list.filter(post_title__icontains=title)
    post_filter = PostFilter(request.GET, queryset=post_list)

    try:
        paginator = Paginator(post_filter.qs, 20)
    except:
        paginator = Paginator(post_filter.qs, 0)

    page = request.GET.get('page')

    try:
        dataqs = paginator.page(page)
    except PageNotAnInteger:
        dataqs = paginator.page(1)
    except EmptyPage:
        dataqs = paginator.page(paginator.num_pages)
    print("testkanna",post_filter,dataqs,title)
    context = {     
        'filter': post_filter,
        'dataqs':dataqs,
        'title':title,
      
    }
    return render(request,'admin/custom/school/post/post_list.html',context)

#Resource Management
@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def resource_list(request):
    resources = Resource.objects.filter(subject_assignment__programme__department__degree_category=request.user.college.category)
    context={     
        'resources':resources,
    }
    return render(request,'admin/custom/school/resource/resource_list.html',context)

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def student_datail(request,pk):
    student= Student.objects.get(id=pk)
    all_quiz= Quiz.objects.filter(subject_assignment__programme__department__degree_category=request.user.college.category,subject_assignment__programme=student.programme,subject_assignment__year=student.year,subject_assignment__semester=student.semester).distinct()
    score=list()
    test_names=list()
    assignment_names=list()
    assignment_score =list()
    test_subjects = list()
    assignment_subjects = []

    for quiz in all_quiz:            
        test_names.append(quiz.name) 
        get_subjects=quiz.subject_assignment.all()
        subject_name=''
        for subject in get_subjects:
            subject_name+=subject.subject.subject_title + ' ,'                
        test_subjects.append(subject_name)

        try:
            attempt=Attempt.objects.get(user=student.user,Quiz=quiz)  
            score.append(attempt.score)
        except:
            score.append('Not Attended')
                
    #get assignment mark
    subject_assignment=Subject_Assignment.objects.filter(programme=student.programme,year=student.year,semester=student.semester)
    assignments=Assignment.objects.filter(topic__croom__subject_assignment__in=subject_assignment,topic__croom__subject_assignment__programme__department__degree_category=request.user.college.category,topic__croom__subject_assignment__year=student.year,topic__croom__subject_assignment__semester=student.semester) 
    for assignment in assignments:
        assignment_names.append(assignment.title) 

        assignment_subjects.append(assignment.topic.croom.subject_assignment.subject.subject_title)
        try:
            submission=Submission.objects.get(student=student,assignment=assignment,status="A")
            assignment_score.append(submission.marks)
            attended+=1
        except:
            assignment_score.append("Not Submitted")

    context={
        'student':student,
        'test_names': test_names,
        'score':score,
        'total_test':all_quiz.count(),
        'assignment_score':assignment_score,
        'assignment_names':assignment_names,
        'total_assignment':assignments.count(),
        'test_subjects':test_subjects,
        'assignment_subjects':assignment_subjects,

    }
    return render(request,'admin/custom/school/student_detail.html',context)
@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def password_reset(request,pk):
    data = dict()
    try:
        user = User.objects.get(Q(student__programme__department__degree_category=request.user.college.category) | Q(faculty__category=request.user.college.category),id=pk)
        new_password = str(random.randint(1000,9999))
        user.password = make_password(new_password)
        user.save()
        data['new_password'] = new_password
        data['valid'] = True
        data['html'] = render_to_string('admin/custom/school/password_show.html',{'new_password':new_password,'user':user,},request)
    except:
        data['valid'] = False
        data['response'] = " Access Denied"

    return JsonResponse(data)

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def user_edit(request,role,pk):
    data = dict()
    if role == "student":
        student = Student.objects.get(id=pk,programme__department__degree_category=request.user.college.category)
        user = User.objects.get(id=student.user.id)
        if request.method == "POST":
            user_form = UserForm(request.POST,instance=user)
            student_form = StudentEditForm(request.POST,instance=student)

            if user_form.is_valid() and student_form.is_valid():
                user_form.save()
                student_form.save()
                student =  Student.objects.get(id=student.id)
                data['valid'] = True
                data['html'] = render_to_string('admin/custom/school/student_data.html',{'student':student,},request)

            else:
                data['valid'] = False
                data['form'] = render_to_string('admin/custom/school/student_edit_profile_form.html',{'user_form':user_form,'student_form':student_form,'pk':pk,},request)


        else:
            user_form = UserForm(instance=user)
            student_form = StudentEditForm(instance=student)
            data['valid'] = True
            data['form'] = render_to_string('admin/custom/school/student_edit_profile_form.html',{'user_form':user_form,'student_form':student_form,'pk':pk,},request)
    
    elif role == 'teacher':
        teacher = Faculty.objects.get(id=pk,category=request.user.college.category)
        user = User.objects.get(id=teacher.user.id)
        if request.method == "POST":
            user_form = UserForm(request.POST,instance=user)
            teacher_form = TeacherEditForm(request.POST,instance=teacher)

            if user_form.is_valid() and teacher_form.is_valid():
                user_form.save()
                teacher_form.save()
                teacher =  Faculty.objects.get(id=teacher.id)
                data['valid'] = True
                data['html'] = render_to_string('admin/custom/school/teacher_data.html',{'teacher':teacher,},request)

            else:
                data['valid'] = False
                data['form'] = render_to_string('admin/custom/school/teacher_edit_profile_form.html',{'user_form':user_form,'teacher_form':teacher_form,'pk':pk,'teacher':teacher,},request)


        else:
            user_form = UserForm(instance=user)
            teacher_form = TeacherEditForm(instance=teacher)
            data['valid'] = True
            data['form'] = render_to_string('admin/custom/school/teacher_edit_profile_form.html',{'user_form':user_form,'teacher_form':teacher_form,'pk':pk,'teacher':teacher,},request)
    

    return JsonResponse(data)

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def view_teacher(request,pk):
    teacher = Faculty.objects.get(id=pk,category=request.user.college.category)
    programme = Programme.objects.filter(department__degree_category = request.user.college.category)
    subject =   Subject.objects.filter(degree_categaroy = request.user.college.category )
    subject_assignment = Subject_Assignment.objects.filter(programme=teacher.programme,programme__department__degree_category=request.user.college.category)
    print(subject_assignment)
    facult_subject_assignment = Faculty_Subject_Assignment.objects.filter(faculty = teacher,subject__in=subject_assignment)
    context = {
        'teacher':teacher,
        'programme':programme,
        'year_options':year_options,
        'subject' :subject,
        'subject_assignment':subject_assignment,
        'semester':semester,
         'section_options':section_options,
        'facult_subject_assignment':facult_subject_assignment
    }
    return render(request,'admin/custom/school/view_teacher.html',context)

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def logbook_history(request,pk):
    teacher = Faculty.objects.get(id=pk,category=request.user.college.category)
    teacher_subjects = Faculty_Subject_Assignment.objects.filter(faculty__category=request.user.college.category,faculty__programme=teacher.programme)
    counter = 1
    tab = request.GET.get('tab')
    if not tab:
        tab = 1
    tab = int(tab)
    for subject in teacher_subjects:        
        if counter == tab:
            subject_instance = subject
        counter += 1
    
    log_book_entries = Log_Book_Entry.objects.filter(log_book__subject=subject_instance )
    context = {
      'log_book_entries':log_book_entries, 
      'tab':tab, 
      'subjects': teacher_subjects,
      'subject_instance':subject_instance,
      'teacher':teacher,
    }
    return  render(request,'admin/custom/school/teacher/logbook-history.html',context)

def subject_assignment(request,pk):
    data = dict()
    faculty = Faculty.objects.filter(id = pk)
    if request.method == "POST":
        section = request.POST.get('section')
        subject = request.POST.get('subject')
        
    if Faculty_Subject_Assignment.objects.filter(subject__id=subject,faculty__id=pk).exists():
        data['valid'] = False
        data['response'] = "Already Added..!"   
    else:
        teacher= Faculty.objects.get(id=pk,category=request.user.college.category)
        subject = Subject_Assignment.objects.get(id=subject)
        subject_assignment = Subject_Assignment.objects.filter(programme=teacher.programme,programme__department__degree_category=request.user.college.category)
        Faculty_Subject_Assignment.objects.create(faculty=teacher,subject=subject,section=section)
        ClassRoom.objects.create(name=subject.subject.subject_title,subject_assignment=subject,created_by=teacher.user,activate=True)        
        facult_subject_assignment = Faculty_Subject_Assignment.objects.filter(faculty = teacher,subject__id__in=subject_assignment)
        print("testing buddy ",subject,teacher,subject_assignment)
        data['valid'] = True
        data['html'] =render_to_string('admin/custom/school/teacher_subject.html',{'teacher':teacher,'facult_subject_assignment':facult_subject_assignment},request)

    return JsonResponse(data)

def subject_report(request,pk,subject):
    data = dict()
    teacher_object = Faculty.objects.get(id=pk)
    subject = Subject_Assignment.objects.get(id=subject)    
    assignments = Assignment.objects.filter(topic__croom__subject_assignment__id=subject.id,topic__croom__subject_assignment__programme__department__degree_category=request.user.college.category)
    assignment_not_submitted_percentage = []
    assignment_pass_percentage_list = []
    assignment_top_score_list = []
    assignment_avg_score_list = []
    assignment_name = []
    assignment_count  = 0
    materials = Material.objects.filter(topic__croom__created_by=teacher_object.user,topic__croom__subject_assignment__programme__department__degree_category=request.user.college.category,topic__croom__subject_assignment=subject)
    discussions = Forum.objects.filter(topic__croom__created_by=teacher_object.user,topic__croom__subject_assignment=subject,topic__croom__subject_assignment__programme__department__degree_category=request.user.college.category)

    for assignment in assignments:
        assignment_count += 1
        assignment_name.append(str(assignment))
        ts = assignment.strength        
        not_sub_count =ts - Submission.objects.filter(assignment=assignment).count()   
        print(not_sub_count)
        assignment_not_submitted_percentage.append(int(float((not_sub_count / ts)) * 100) )
        pp_count = Submission.objects.filter(assignment=assignment,marks__gte=50).count()
        pp = (float(pp_count / ts)) * 100
        assignment_pass_percentage_list.append(int(pp))
        top_score = Submission.objects.filter(assignment=assignment,marks__gte=50).aggregate(Max('marks'))
        avg_score = Submission.objects.filter(assignment=assignment,marks__gte=50).aggregate(Avg('marks'))
        if top_score['marks__max'] is not None:               
            top_mark = top_score['marks__max']                
        else:
            top_mark = 0

        if avg_score['marks__avg'] is not None:               
            avg_mark = avg_score['marks__avg']                
        else:
            avg_mark = 0

        assignment_top_score_list.append(int(top_mark))
        assignment_avg_score_list.append(int(avg_mark))


    tests = Quiz.objects.filter(subject_assignment__programme__department__degree_category=request.user.college.category,subject_assignment=subject)
    test_not_submitted_percentage = []
    test_pass_percentage_list = []
    test_top_score_list = []
    test_avg_score_list = []
    test_name_list = []
    test_count = 0
    test_id = []
    open_date = []
    close_date = []
    timed_exam = []
    duration  = []
    total_strength = []
    attempt = []
    not_attempt = []  
    subject_id = []  
    for test in tests:
        subject_id.append(subject.id)
        test_id.append(test.id)      
        open_date.append(str(test.openDate))
        close_date.append(str(test.closeDate))
        if test.timed_exam:
            duration.append(test.duration_in_minutes)
        else:
            duration.append('NA')
        test_count += 1
        test_name_list.append(str(test))
        students_count = Student.objects.filter(grade=subject.grade,group=subject.group,section=section,school=request.user.school).count()
        total_strength.append(students_count)
        pp_count = Attempt.objects.filter(Quiz=test,user__student__section=section,user__student__grade=subject.grade,user__student__group=subject.group,user__student__school=request.user.school,score__gte=50).count()
        attempt_count = Attempt.objects.filter(Quiz=test,user__student__section=section,user__student__grade=subject.grade,user__student__group=subject.group,user__student__school=request.user.school).count()
        attempt.append(attempt_count)
        pp = (pp_count / students_count) * 100      
        not_atempt_count =students_count - attempt_count  
        not_attempt.append(not_atempt_count) 
        test_not_submitted_percentage.append(int(float((not_atempt_count / students_count)) * 100) )
        test_pass_percentage_list.append(pp)
        top_score = Attempt.objects.filter(Quiz=test,user__student__section=section,user__student__grade=subject.grade,user__student__group=subject.group,user__student__school=request.user.school).aggregate(Max('score'))
        avg_score = Attempt.objects.filter(Quiz=test,user__student__section=section,user__student__grade=subject.grade,user__student__group=subject.group,user__student__school=request.user.school).aggregate(Avg('score'))
        if top_score['score__max'] is not None:               
            top_mark = top_score['score__max']                
        else:
            top_mark = 0 
        if avg_score['score__avg'] is not None:               
            avg_mark = avg_score['score__avg']                
        else:
            avg_mark = 0
        test_top_score_list.append(int(top_mark))
        test_avg_score_list.append(int(avg_mark))

    zippedList=[]
    zippedList= zip(test_id,test_name_list,open_date,close_date,duration,total_strength,attempt,not_attempt)
    

    context = {
        'subject':subject,
        'materials':materials,
        'discussions':discussions,
        'teacher':teacher_object,
        #'grade':subject.grade.grade,
        #'section':section,
        'assignments':assignments,
        'zippedList':zippedList,
        'test_name_list':test_name_list,
        'test_pass_percentage_list':test_pass_percentage_list,
        'test_not_submitted_percentage':test_not_submitted_percentage,
        'test_top_score_list':test_top_score_list,
        'test_avg_score_list':test_avg_score_list,
        'assignment_name':assignment_name,
        'assignment_avg_score_list':assignment_avg_score_list,
        'assignment_top_score_list':assignment_top_score_list,
        'assignment_pass_percentage_list':assignment_pass_percentage_list,
        'assignment_not_submitted_percentage':assignment_not_submitted_percentage,

    }

    return render(request,'admin/custom/school/teacher/subject_report.html',context)

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def remove_subject(request,pk):
    data = dict()
    print(pk)
    subject = Faculty_Subject_Assignment.objects.get(faculty__id=pk)

    if subject.faculty.category == request.user.college.category:
        print("hell9euw")
        subject.delete()
        data['valid'] =  True
        data['pk'] = pk
    else:
        data['valid'] =  False

    return JsonResponse(data)

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def post_add(request):
    data = dict()
    category = request.user.college.category
    if request.method == "POST":
        form = PostForm(request.POST,request=request)
        if form.is_valid():
            post = form.save(commit=False)
            post.degree_categaroy = category
            post.posted_by = request.user
            post.save()
            form.save_m2m()
            data['html'] = render_to_string('admin/custom/school/post/post.html',{'post':post},request)
            data['valid'] = True
        else:
            context = {
                'form':form,
                'operation':'Add',
            }
            data['form'] = render_to_string('admin/custom/school/post/form.html',context,request)
            data['valid'] = False

    else:
        form = PostForm(request=request)
        context = {
            'form':form,
            'operation':'Add',
        }        
        data['form'] = render_to_string('admin/custom/school/post/form.html',context,request)
        data['valid'] = True
    return JsonResponse(data) 

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def video_list(request,pk):
    print("id",pk)
    resource = Resource.objects.get(id=pk)
    print("test-resource",resource)
    videos = Video.objects.filter(resource=resource).order_by('-id')
    context={     
        'resource':resource,
        'videos':videos,
    }
    return render(request,'admin/custom/school/resource/video_list.html',context)

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def add_video(request,pk):
    resource = Resource.objects.get(id=pk,subject_assignment__programme__department__degree_category=request.user.college.category)
    data = dict()
    if request.method == "POST":
        form = VideoForm(request.POST)
        if form.is_valid():
            video = form.save(commit=False)
            video.resource=resource
            video.save()
            videos = Video.objects.filter(resource=resource).order_by('-id')
            data['html'] = render_to_string('admin/custom/school/resource/video.html',{'videos':videos},request)
            data['valid'] = True

        else:
            context = {
                'form':form,
                'operation':'Add',
                'pk':pk,
            }
            data['form'] = render_to_string('admin/custom/school/resource/video_form.html',context,request)
            data['valid'] = False

    else:
        form = VideoForm()
        context = {
            'form':form,
            'operation':'Add',
            'pk':pk,
        }        
        data['form'] = render_to_string('admin/custom/school/resource/video_form.html',context,request)
        data['valid'] = True

    return JsonResponse(data) 


@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def edit_video(request,pk):
    data = dict()
    video = Video.objects.get(id=pk,resource__subject_assignment__programme__department__degree_category=request.user.college.category)
    if request.method == "POST":
        form = VideoForm(request.POST,instance=video)
        if form.is_valid():
            form.save()
            videos = Video.objects.filter(resource=video.resource).order_by('-id')
            data['html'] = render_to_string('admin/custom/school/resource/video.html',{'videos':videos},request)
            data['valid'] = True
        else:
            context = {
                'form':form,
                'operation':'Update',
                'pk':pk,
            }
            
            data['form'] = render_to_string('admin/custom/school/resource/video_form.html',context,request)
            data['valid'] = False

    else:
        form = VideoForm(instance=video)
        context = {
            'form':form,
            'operation':'Update',
            'pk':pk,
        }        
        data['form'] = render_to_string('admin/custom/school/resource/video_form.html',context,request)
        data['valid'] = True

    return JsonResponse(data) 

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def video_delete(request,pk):
    data = dict()
    video = Video.objects.get(id=pk,resource__subject_assignment__programme__department__degree_category=request.user.college.category)
    resource =video.resource 
    if video.resource.subject_assignment.programme.department.degree_category  == request.user.college.category:
        video.delete()  
        videos = Video.objects.filter(resource=resource).order_by('-id')
        data['html'] = render_to_string('admin/custom/school/resource/video.html',{'videos':videos},request)
        data['valid'] = True
    else:
        data['valid'] = False
    return JsonResponse(data) 


@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def view_class_report(request,grade,year):

    subject_ids =  Resource.objects.filter(subject_assignment__programme=grade,subject_assignment__programme__department__degree_category=request.user.college.category).values_list('subject_assignment__id').distinct()
    subjects = Subject_Assignment.objects.filter(programme=grade,programme__department__degree_category=request.user.college.category,year=year).distinct()  
    assignment_not_submitted_percentage = []
    assignment_pass_percentage_list = []
    assignment_top_score_list = []
    assignment_avg_score_list = []
    assignment_name = []    
    test_not_submitted_percentage = []
    test_pass_percentage_list = []
    test_top_score_list = []
    test_avg_score_list = []
    test_name_list = []
    test_count = 0
    test_id = []
    subject_id = []
    open_date = []
    close_date = []
    timed_exam = []
    duration  = []
    total_strength = []
    attempt = []
    not_attempt = []
    assignments = []

    for subject in subjects:
        assignment = Assignment.objects.filter(topic__croom__subject_assignment=subject.id,topic__croom__subject_assignment__programme__department__degree_category=request.user.college.category).last()
        if assignment:  
            assignments.append(assignment)          
            assignment_name.append(str(assignment))
            ts = assignment.strength  
            if ts == 0:
                assignment_not_submitted_percentage.append(0)
                assignment_pass_percentage_list.append(0)
                assignment_top_score_list.append(0)
                assignment_avg_score_list.append(0)
            else:
                        
                not_sub_count =ts - Submission.objects.filter(assignment=assignment).count()   
                assignment_not_submitted_percentage.append(int(float((not_sub_count / ts)) * 100) )
                pp_count = Submission.objects.filter(assignment=assignment,marks__gte=50).count()
                pp = (float(pp_count / ts)) * 100
                assignment_pass_percentage_list.append(int(pp))
                top_score = Submission.objects.filter(assignment=assignment,marks__gte=50).aggregate(Max('marks'))
                avg_score = Submission.objects.filter(assignment=assignment,marks__gte=50).aggregate(Avg('marks'))
                if top_score['marks__max'] is not None:               
                    top_mark = top_score['marks__max']                
                else:
                    top_mark = 0

                if avg_score['marks__avg'] is not None:               
                    avg_mark = avg_score['marks__avg']                
                else:
                    avg_mark = 0

                assignment_top_score_list.append(int(top_mark))
                assignment_avg_score_list.append(int(avg_mark))

        test = Quiz.objects.filter(subject_assignment=subject.id,subject_assignment__programme__department__degree_category=request.user.college.category,subject_assignment__programme=grade,isActive=True,openDate__lte=date.today()).last()
        if test:  
            subject_id.append(subject.id)           
            test_id.append(test.id)      
            open_date.append(str(test.openDate))
            close_date.append(str(test.closeDate))
            if test.timed_exam:
                duration.append(test.duration_in_minutes)
            else:
                duration.append('NA')
            test_count += 1
            test_name_list.append(str(test))
            students_count = Student.objects.filter(programme=subject.programme,year=year,programme__department__degree_category=request.user.college.category).count()
            if students_count == 0:
                total_strength.append(0)
                attempt.append(0)
                not_attempt.append(0) 
                test_not_submitted_percentage.append(0)
                test_pass_percentage_list.append(0)    
                test_top_score_list.append(0)
                test_avg_score_list.append(0)

            else:

                total_strength.append(students_count)          
                pp_count = Attempt.objects.filter(Quiz=test,user__student__year=year,user__student__programme=subject.programme,user__student__programme__department__degree_category=request.user.college.category,score__gte=50).count()
                attempt_count = Attempt.objects.filter(Quiz=test,user__student__year=year,user__student__programme=subject.programme,user__student__programme__department__degree_category=request.user.college.category).count()
                attempt.append(attempt_count)                
                pp = (pp_count / students_count) * 100                      
                not_atempt_count =students_count - attempt_count  
                not_attempt.append(not_atempt_count) 
                test_not_submitted_percentage.append(int(float((not_atempt_count / students_count)) * 100) )
                test_pass_percentage_list.append(pp)
                top_score = Attempt.objects.filter(Quiz=test,user__student__year=year,user__student__programme=subject.programme,user__student__programme__department__degree_category=request.user.college.category).aggregate(Max('score'))
                avg_score = Attempt.objects.filter(Quiz=test,user__student__year=year,user__student__programme=subject.programme,user__student__programme__department__degree_category=request.user.college.category).aggregate(Avg('score'))
                if top_score['score__max'] is not None:               
                    top_mark = top_score['score__max']                
                else:
                    top_mark = 0 

                if avg_score['score__avg'] is not None:               
                    avg_mark = avg_score['score__avg']                
                else:
                    avg_mark = 0

                test_top_score_list.append(int(top_mark))
                test_avg_score_list.append(int(avg_mark))

        zippedList=[]
        zippedList = zip(test_id,subject_id,test_name_list,open_date,close_date,duration,total_strength,attempt,not_attempt)
        programme = Programme.objects.get(id=grade)
    #groups = gmGroup.objects.all()
    context = {
        'subjects':subjects,
        'zippedList':zippedList,
        'grade':grade,
        'section':year,
        'test_name_list':test_name_list,
        'test_not_submitted_percentage':test_not_submitted_percentage,
        'test_pass_percentage_list':test_pass_percentage_list,
        'test_top_score_list':test_top_score_list,
        'test_avg_score_list':test_avg_score_list,
        'assignment_not_submitted_percentage':assignment_not_submitted_percentage,
        'assignment_pass_percentage_list':assignment_pass_percentage_list,
        'assignment_top_score_list':assignment_top_score_list,
        'assignment_avg_score_list':assignment_avg_score_list,
        'assignment_name':assignment_name,
        'assignments':assignments,
        'programme':programme 
        
    }
    return render(request,'admin/custom/school/performance/performance.html',context)

def get_subject_report(request,subject,section):
    data = dict()
    subject = Subject_Assignment.objects.get(id=subject)    
    assignments = Assignment.objects.filter(topic__croom__subject_assignment__id=subject.id,topic__croom__subject_assignment__programme__department__degree_category=request.user.college.category)
    print("test assignment",assignments)
    assignment_not_submitted_percentage = []
    assignment_pass_percentage_list = []
    assignment_top_score_list = []
    assignment_avg_score_list = []
    assignment_name = []
    assignment_count  = 0
    materials = Material.objects.filter(topic__croom__subject_assignment__programme__department__degree_category=request.user.college.category,topic__croom__subject_assignment=subject,topic__croom__subject_assignment__year=section)
    discussions = Forum.objects.filter(topic__croom__subject_assignment=subject,topic__croom__subject_assignment__programme__department__degree_category=request.user.college.category)

    for assignment in assignments:
        assignment_count += 1
        assignment_name.append(str(assignment))
        ts = assignment.strength 
        if ts == 0:
            assignment_not_submitted_percentage.append(0)
            assignment_pass_percentage_list.append(0)
            assignment_top_score_list.append(0)
            assignment_avg_score_list.append(0)
        else:                          
            not_sub_count =ts - Submission.objects.filter(assignment=assignment).count()   
            assignment_not_submitted_percentage.append(int(float((not_sub_count / ts)) * 100) )
            pp_count = Submission.objects.filter(assignment=assignment,marks__gte=50).count()
            pp = (float(pp_count / ts)) * 100
            assignment_pass_percentage_list.append(int(pp))
            top_score = Submission.objects.filter(assignment=assignment,marks__gte=50).aggregate(Max('marks'))
            avg_score = Submission.objects.filter(assignment=assignment,marks__gte=50).aggregate(Avg('marks'))
            if top_score['marks__max'] is not None:               
                top_mark = top_score['marks__max']                
            else:
                top_mark = 0

            if avg_score['marks__avg'] is not None:               
                avg_mark = avg_score['marks__avg']                
            else:
                avg_mark = 0

            assignment_top_score_list.append(int(top_mark))
            assignment_avg_score_list.append(int(avg_mark))


    tests = Quiz.objects.filter(subject_assignment__programme__department__degree_category=request.user.college.category,subject_assignment=subject)
    test_not_submitted_percentage = []
    test_pass_percentage_list = []
    test_top_score_list = []
    test_avg_score_list = []
    test_name_list = []
    test_count = 0
    test_id = []
    open_date = []
    close_date = []
    timed_exam = []
    duration  = []
    total_strength = []
    attempt = []
    not_attempt = []    
    for test in tests:
        test_id.append(test.id)      
        open_date.append(str(test.openDate))
        close_date.append(str(test.closeDate))
        if test.timed_exam:
            duration.append(test.duration_in_minutes)
        else:
            duration.append('NA')
        test_count += 1
        test_name_list.append(str(test))
        students_count = Student.objects.filter(programme=subject.programme,year=section,programme__department__degree_category=request.user.college.category).count()
        if students_count == 0:
            total_strength.append(0)
            attempt.append(0)
            not_attempt.append(0) 
            test_not_submitted_percentage.append(0)
            test_pass_percentage_list.append(0)    
            test_top_score_list.append(0)
            test_avg_score_list.append(0)
        else:
            total_strength.append(students_count)
            pp_count = Attempt.objects.filter(Quiz=test,user__student__year=section,user__student__programme__department__degree_category=subject.programme.department.degree_category,score__gte=50).count()
            attempt_count = Attempt.objects.filter(Quiz=test,user__student__year=section,user__student__programme=subject.programme,user__student__programme__department__degree_category=request.user.college.category).count()
            attempt.append(attempt_count)
            pp = (pp_count / students_count) * 100      
            not_atempt_count =students_count - attempt_count  
            not_attempt.append(not_atempt_count) 
            test_not_submitted_percentage.append(int(float((not_atempt_count / students_count)) * 100) )
            test_pass_percentage_list.append(pp)
            top_score = Attempt.objects.filter(Quiz=test,user__student__year=section,user__student__programme=subject.programme,user__student__programme__department__degree_category=request.user.college.category).aggregate(Max('score'))
            avg_score = Attempt.objects.filter(Quiz=test,user__student__year=section,user__student__programme=subject.programme,user__student__programme__department__degree_category=request.user.college.category).aggregate(Avg('score'))
            if top_score['score__max'] is not None:               
                top_mark = top_score['score__max']                
            else:
                top_mark = 0 
            if avg_score['score__avg'] is not None:               
                avg_mark = avg_score['score__avg']                
            else:
                avg_mark = 0
            test_top_score_list.append(int(top_mark))
            test_avg_score_list.append(int(avg_mark))

    zippedList=[]
    zippedList = zip(test_id,test_name_list,open_date,close_date,duration,total_strength,attempt,not_attempt)
    

    data['test'] = test_name_list
    data['test_pass_percentage_list'] = test_pass_percentage_list
    data['test_not_submitted_percentage'] = test_not_submitted_percentage
    data['test_top_score_list'] = test_top_score_list
    data['test_avg_score_list'] = test_avg_score_list

    data['assignment_not_submitted_percentage'] = assignment_not_submitted_percentage
    data['assignment_pass_percentage_list'] = assignment_pass_percentage_list
    data['assignment_top_score_list'] = assignment_top_score_list
    data['assignment_avg_score_list'] = assignment_avg_score_list
    data['assignment_name'] = assignment_name
    print("assignments",assignments)
    data['assignments'] = render_to_string('admin/custom/school/score/assignment_list.html',{'assignments':assignments},request)
    data['tests'] = render_to_string('admin/custom/school/score/test_list.html',{'zippedList':zippedList,'subject':subject,'grade':subject.programme,'section':section,},request)
    data['materials'] = render_to_string('admin/custom/school/performance/material_list.html',{'materials':materials,},request)
    data['discussions'] = render_to_string('admin/custom/school/performance/discussions_list.html',{'discussions':discussions,},request)
    return JsonResponse(data)


@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def get_test_report(request,pk,subject,grade,section):
    data =  dict()
    pass_percentage_list = []
    top_score_list = []
    avg_score_list = []
    test_name_list = [] 
    not_attempt_percentage = []  
    test_object = Quiz.objects.get(id=pk)
    #test_obj = Quiz.objects.filter(id=pk)
    #for test_obj in test_obj.subject_assignment.all:
        #print("test",test_object,test_obj.programme)
    if test_object:
        subject = Subject_Assignment.objects.get(id=subject)
        students_count = Student.objects.filter(programme=subject.programme).count()
        tests = Quiz.objects.filter(subject_assignment__programme__department__degree_category=request.user.college.category)
        print("attempted list",test_object,grade)
        attempted_list=Attempt.objects.filter(Quiz=test_object,Quiz__subject_assignment__programme = grade)
        not_attempted_students  = Student.objects.filter(programme = subject.programme,programme__department__degree_category=subject.programme.department.degree_category,year = subject.year).exclude(user__id__in=Attempt.objects.filter(Quiz=test_object).values_list('user'))
        data['html'] =  render_to_string('admin/custom/school/score/test_report.html',{'grade':subject.programme,'subject':subject.id,'section':section, 'attempted_list':attempted_list,'test':test_object,'not_attempted_students':not_attempted_students,},request)
        return JsonResponse(data)
    else:
        return JsonResponse('Access Denied')
    
    
@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def get_assignment_report(request,pk):
    data =  dict()
    pass_percentage_list = []
    top_score_list = []
    avg_score_list = []
    assignment_name_list = []
    assignment_object = Assignment.objects.get(id=pk,topic__croom__subject_assignment__programme__department__degree_category=request.user.college.category)
    submissions = Submission.objects.filter(assignment=assignment_object) 
    subject_assignment=Subject_Assignment.objects.get(id=assignment_object.topic.croom.subject_assignment.id)
    not_submitted_students = Student.objects.filter(programme__department__degree_category=request.user.college.category,programme=subject_assignment.programme).exclude(id__in=Submission.objects.filter(assignment_id=assignment_object).values_list('student'))
    data['html'] =  render_to_string('admin/custom/school/score/assignment_report.html',{'assignment':assignment_object,'submissions':submissions,'not_submitted_students':not_submitted_students,},request)
    return JsonResponse(data)

def subject_load(request):
    data = dict()
    group = request.GET.get('group')
    grade = request.GET.get('grade')
    section = request.GET.get('section') 
    if group == "All":
        subjects = Subject_Assignment.objects.filter(grade__grade=grade,board=request.user.school.board)        
    else:
        subjects = Subject_Assignment.objects.filter(group__id=group,grade__grade=grade,board=request.user.school.board)
        
    data['html'] =  render_to_string('admin/custom/school/performance/subject-list.html',{'subjects':subjects,'section':section},request)
    return JsonResponse(data)

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def post_edit(request,pk):
    data = dict()
    print(pk,request.user.college.category)
    post = Post.objects.get(id=pk)
    print(post)
    if request.method == "POST":
        form = PostForm(request.POST,instance=post,request=request)
        if form.is_valid():
            post = form.save(commit=False)
            post.degree_category=request.user.college.category
            post.posted_by = request.user
            post.save()
            form.save_m2m()
            data['html'] = render_to_string('admin/custom/school/post/post_td.html',{'post':post},request)
            data['valid'] = True
            data['pk'] = pk
        else:
            context = {
                'form':form,
                'operation':'Update',
                'pk':pk,
            }
            data['form'] = render_to_string('admin/custom/school/post/form.html',context,request)
            data['valid'] = False

    else:
        form = PostForm(instance=post,request=request)
        context = {
            'form':form,
            'operation':'Update',
            'pk':pk,
        }        
        data['form'] = render_to_string('admin/custom/school/post/form.html',context,request)
        data['valid'] = True

    return JsonResponse(data) 

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def post_delete(request,pk):
    data = dict()
    post = Post.objects.get(id=pk)
    if post.posted_by  == request.user:
        post.delete()  
        data['pk'] = pk
        data['valid'] = True
    else:
        data['valid'] = False

    return JsonResponse(data)
 
@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def teacher_subject_load(request,grade):
    data = dict()
    programme = grade
    print("test-programme",programme)
    #subjects = Subject.objects.filter(degree_categaroy=request.user.college.category)
    subjects = Subject_Assignment.objects.filter(programme__department__degree_category=request.user.college.category,programme=programme)
    print("teachersubject",subjects)
    data['html'] = render_to_string('admin/custom/school/subject.html',{'subjects':subjects,},request)
    data['valid'] = True
    return JsonResponse(data)

def ajax_get_student(request):
    data= dict()
    grade= request.POST.get('grade')
    section= request.POST.get('section')

    if grade:
        if grade and section:
            student=Student.objects.filter(school=request.user.school,grade=grade,section=section)
        else:
            student=Student.objects.filter(school=request.user.school,grade=grade)

    
    data['form'] = render_to_string('admin/custom/school/student.html',{'students':student,},request)

    return JsonResponse(data)

@login_required(login_url='login') #Authentication
@user_passes_test(is_college,login_url='login') #Authorization
def assignment_list(request,subject,grade,section):
    subject = Subject_Assignment.objects.get(id=subject)    
    assignments = Assignment.objects.filter(topic__croom__school=request.user.school,topic__croom__subject=subject)
    context = {
        'assignments':assignments,
        'subject':subject,
        'grade':grade,
        'section':section,
    }
    return render(request,'admin/custom/school/assignment_list.html',context)

class Subject_assignmentFilter(django_filters.FilterSet):
    #programme = django_filters.ModelMultipleChoiceFilter(queryset=Programme.objects.filter(department__degree_category = request.user.college.category))    
    class Meta:
        model = Subject_Assignment
        fields = ['programme','subject','year','semester']

def programme_list(request):
    if "POST" == request.method:
        #subject_assignment = Subject_Assignment()
        #dataset = Dataset()
        error_row = []            
        excel_file = request.FILES["question_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"] 
        excel_data = list()
         
        for count, row in enumerate(worksheet.iter_rows()):               
            if count == 0:
                continue
            elif row[0].value == 'end':
                break
            
            title = str(row[0].value).replace('\n','<br>')
            department = str(row[1].value).replace('\n','<br>')
            graduation = str(row[2].value).replace('\n','<br>')
            subject = str(row[3].value).replace('\n','<br>')
            year = str(row[4].value).replace('\n','<br>')
            semester = str(row[5].value).replace('\n','<br>')
            print("test",title,department,graduation,subject,year,semester)
            program_instance = Programme.objects.get(title=title,department__name=department,graduation__description=graduation)
            subject_instance = Subject.objects.get(subject_title=subject)
            print(program_instance.id,subject_instance.id)
            Subject_Assignment.objects.create(programme=program_instance,subject=subject_instance,year=year,semester=semester)
            
        '''
            if row[0].value == None or  row[4].value == None or  row[1].value == None:
                error_row.append("row " + str(count+1))
                continue
            
            for title in title:
                print("test",title,department,graduation,subject,year,semester)
                programe_id = Programme.objects.get(icontains=prog)
                print("programee",programe_id)
            
            print("test count ",programme)
            #create subject by excel
            degree_category = request.user.college.category
            degree_category = Degree_Categaroy.objects.filter(category=degree_category)
            for degree_category in degree_category:
                degree_category = degree_category.id
            print(degree_category)
            subject_category = Subject_Category.objects.filter(category_name=subject_category)

           
            for sub in subject_category:
                created_subject = Subject.objects.create(subject_title=subject,subject_category_id= sub.id,degree_categaroy_id=degree_category)
                print("subject",created_subject)
                subject_id = Subject.objects.filter(subject_title=subject)
                print("subject_id",subject_id)
                #programme_id = Programme.objects.filter(id=programme)
                print("programme instance",programme)
                Subject_Assignment.objects.create(programme__in=programme,subject=sub,year=year,semester=semester)

            if row[0].value == None or  row[5].value == None or  row[1].value == None:
                error_row.append("row " + str(count+1))
                continue
             '''   
            #Subject_Assignment.objects.create(programme=programme,subject=subject,year=year,semester=semester)
        #imported_data = dataset.load(excel_file.read(), format='csv', headers=False,encoding = "utf-8")
        #print(imported_data)
        #for data in imported_data:
            #print(data)
        #result = subject_assignment.import_data(dataset, dry_run=True)  # Test the data import

        #if not result.has_errors():
            #subject_assignment.import_data(dataset, dry_run=False)  
    else :
        subject_assignment = Subject_Assignment.objects.filter(programme__department__degree_category=request.user.college.category)
        subject_assignmentfilter = Subject_assignmentFilter(request.GET, queryset=subject_assignment)

        paginator = Paginator(subject_assignmentfilter.qs, 20)
        page = request.GET.get('page')
        try:
            dataqs = paginator.page(page)
        except PageNotAnInteger:
            dataqs = paginator.page(1)
        except EmptyPage:
            dataqs = paginator.page(paginator.num_pages)
        context ={
            'subject_assignment':subject_assignmentfilter,
            'dataqs':dataqs
        }
    return render(request,'admin/custom/school/programme/programme_list.html',context)

def create_chapter(request,id):
    data = dict()
    if "POST" == request.method:
        subject_assignment = Subject_Assignment.objects.get(id=id)
        #dataset = Dataset()
        error_row = []            
        excel_file = request.FILES["question_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"] 
        excel_data = list()
         
        for count, row in enumerate(worksheet.iter_rows()):               
            if count == 0:
                continue
            elif row[0].value == 'end':
                break
            
            title = str(row[0].value).replace('\n','<br>')
            print("chapter",title)
            Chapter.objects.create(subject_assignment=subject_assignment,name=title)
        chapters = Chapter.objects.filter(id=id)
        data['output'] =True
        data['id'] = id
        #data['output'] = render_to_string('admin/custom/school/programme/chapter_string.html',{'chapters':chapters,'id':id,},request)
        return JsonResponse(data)
        





def view_chapters(request,id):
        chapters = Chapter.objects.filter(subject_assignment__id=id)
        context ={
            'chapters':chapters,
            'id':id
        }
        return render(request,'admin/custom/school/programme/create_chapter.html',context)